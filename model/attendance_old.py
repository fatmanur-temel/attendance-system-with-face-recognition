from datetime import date
from datetime import datetime

import cv2
import openpyxl
import pandas as pd


class Attendance:
    def __int__(self, study_name):
        self.study_name = study_name

    def create_file(self):
        today = date.today()
        day = today.strftime("%b-%d-%Y")
        file_str = self.study_name + "-" + day + ".csv"

        file = open(file_str, "a")
        file.write("Numara, Ad, Soyad, Saat")
        file.close()

        return file_str

    @staticmethod
    def create_attendance(name, file_name, frame):
        with open(file_name, 'r+') as f:
            my_data_list = f.readlines()
            image_name = name.split(',')
            for line in my_data_list:
                entry = line.split(',')
                if image_name[0] == entry[0]:
                    return

            #if name not in file:
            now = datetime.now()
            dt_string = now.strftime('%H:%M:%S')
            f.writelines(f'\n{name},{dt_string}')
            cv2.putText(frame, "Kayit alindi.", (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 0, 0), 3)
            cv2.imshow("Frame", frame)
            cv2.waitKey(500)

    @staticmethod
    def create_excel(file_name):
        data = pd.read_csv(file_name)
        wb = openpyxl.Workbook()
        page = wb.active

        total_rows = len(data)  ### toplan satır sayısı
        total_columns = len(data.columns)  ### toplam sütun sayısı
        print('satır uzunluğu: ', total_rows)
        print('sütun sayısı: ', total_columns)

        for title in range(total_columns):  ### sütun başlıklarını yazdırma döngüsü
            c = title + 1
            page.cell(row=1, column=c).value = data.columns[title]

        for line in range(total_rows):  ### tüm satırlardaki verileri excele yazdırma döngüsü
            for clm in range(total_columns):
                r = line + 2
                c = clm + 1
                page.cell(row=r, column=c).value = data.iat[line, clm]

        file_name = file_name.rstrip('.csv')
        wb.save(file_name + ".xlsx")